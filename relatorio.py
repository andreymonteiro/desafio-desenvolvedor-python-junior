from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

CABECALHO = ["Pacote", "Versão", "Licença", "Última publicação", "Score Snyk", "Vulnerabilidades"]

PREENCHIMENTO_RUIM = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def gerar_relatorio(dados, caminho_saida):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dependências"

    ws.append(CABECALHO)
    for cel in ws[1]:
        cel.font = Font(bold=True)

    for item in dados:
        linha = [
            item["nome"],
            item.get("versao") or "-",
            item.get("licenca") or "-",
            item.get("data_publicacao") or "-",
            item.get("score") if item.get("score") is not None else "N/A",
            item.get("vulnerabilidades") if item.get("vulnerabilidades") is not None else "N/A",
        ]
        ws.append(linha)

        score = item.get("score")
        if isinstance(score, int) and score < 65:
            for cel in ws[ws.max_row]:
                cel.fill = PREENCHIMENTO_RUIM

    for coluna in ws.columns:
        maior = max(len(str(c.value)) for c in coluna)
        ws.column_dimensions[coluna[0].column_letter].width = maior + 2

    wb.save(caminho_saida)
